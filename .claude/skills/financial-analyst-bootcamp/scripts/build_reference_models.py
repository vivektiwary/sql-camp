#!/usr/bin/env python3
"""
build_reference_models.py -- build the three reference Excel workbooks.

    python3 scripts/build_reference_models.py [output_dir]

These are the "known good" models a student compares their own build against.
They are generated rather than hand-built for three reasons:

  * every number traces to a query against the sqlcamp database,
  * they can be rebuilt for a different company by changing COMPANY below,
  * the script itself is Module 24 teaching material -- this is what
    "automate the pack" actually looks like.

Historical figures are hardcoded as INPUTS (blue) because that is what they
are: actuals. Every forecast cell is a formula. No forecast number is typed.

Requires: openpyxl. Recalculate afterwards with the xlsx skill's recalc.py,
or just open in Excel.
"""
import sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Source data -- pulled from the sqlcamp database. The query that produced
# each block is quoted above it so the student can re-run and verify.
# ---------------------------------------------------------------------------
COMPANY = dict(ticker="KVRA", name="Kaveri Retail Ltd", sector="Consumer Discretionary",
               shares_m=87.17, price=149.6865, price_date="31-Mar-2026")

# SELECT d.fiscal_year, SUM(revenue), SUM(cogs), ... FROM fs_income_statement i
# JOIN dim_date d ON d.date_key = i.period_end WHERE i.company_id = 1 GROUP BY 1;
HIST_IS = {                     # FY24, FY25, FY26
    "revenue":      [8249, 8671, 9207],
    "cogs":         [4784, 5029, 5340],
    "opex":         [2351, 2471, 2624],
    "dep":          [284, 293, 305],
    "amort":        [8, 7, 7],
    "int_exp":      [312, 288, 265],
    "int_inc":      [82, 83, 86],
    "tax":          [148, 166, 188],
}
# SELECT ... FROM fs_balance_sheet WHERE company_id = 1
#  AND period_end IN ('2024-03-31','2025-03-31','2026-03-31');
HIST_BS = {
    "cash":     [3380, 3532, 3661],  "ar":    [730, 726, 773],
    "inv":      [1058, 1053, 1121],  "oca":   [111, 110, 118],
    "ppe":      [3619, 3754, 3909],  "gi":    [126, 119, 112],
    "oa":       [195, 195, 195],     "ap":    [705, 702, 747],
    "accr":     [89, 88, 94],        "dr":    [111, 110, 118],
    "std":      [292, 292, 292],     "ltd":   [2566, 2343, 2120],
    "ol":       [117, 117, 117],     "sc":    [973, 973, 973],
    "re":       [4366, 4864, 5428],
}
# SELECT d.fiscal_year, SUM(-capex), SUM(-debt_repaid) FROM fs_cash_flow ...
HIST_CF = {"capex": [400, 429, 460], "debt_repaid": [223, 223, 223]}

# WITH f AS (SELECT company_id, SUM(revenue), SUM(ebitda), ... FROM
#   fs_income_statement WHERE period_end > '2025-03-31'
#   AND period_end <= '2026-03-31' GROUP BY 1) SELECT ... ;
# ticker, name, sector, LTM revenue, EBITDA, EBIT, net income,
# shares (m), price, gross debt, cash, LTM revenue growth %
COMPS = [
 ("ZNTH","Zenith Telecom Ltd","Communication Services",13787,3791,2327,1372,88.48,272.6698,5997,4598,6.6),
 ("BHMT","Bhima Motors Ltd","Consumer Discretionary",14186,1915,1388,619,184.60,78.2831,6753,5930,8.2),
 ("KVRA","Kaveri Retail Ltd","Consumer Discretionary",9207,1243,931,564,87.17,149.6865,2412,3661,6.2),
 ("NLSP","Nilgiri Spice Foods Ltd","Consumer Staples",14632,1975,1357,579,136.01,138.0447,6488,3078,6.1),
 ("PLMT","Palmetto Grocers Inc","Consumer Staples",4666,630,430,140,32.96,133.2541,2514,853,6.5),
 ("FRNC","Frontier Energy Corp","Energy",6231,1090,512,223,10.71,222.8097,3221,2851,2.7),
 ("KSHB","Kosambi Housing Finance Ltd","Financials",10773,3178,2799,1896,129.64,182.4815,4510,7153,13.0),
 ("STNM","Stonemark Bancorp","Financials",17547,5177,4488,3377,1016.63,51.7056,3685,16180,8.5),
 ("ATLB","Atlas Biosciences Inc","Health Care",9870,1431,1059,565,89.67,165.8370,3861,3092,13.1),
 ("VNDH","Vindhya Pharma Ltd","Health Care",10421,1511,1085,676,120.47,161.8067,2266,2710,9.5),
 ("HRBR","Harbourline Freight Corp","Industrials",10435,1200,703,477,47.19,196.6435,1750,4555,5.4),
 ("TPTI","Taptee Logistics Ltd","Industrials",6957,800,438,170,26.70,146.6757,2355,1432,4.8),
 ("CLDW","Cloudwell Systems Inc","Information Technology",12309,2769,2470,1684,376.68,160.0288,3356,4851,15.8),
 ("MRDN","Meridian Softworks Ltd","Information Technology",14610,3287,2887,1855,1313.50,44.5829,6298,6007,12.1),
 ("NRTH","Northgate Analytics Inc","Information Technology",3249,731,634,465,94.34,154.9136,707,1976,14.6),
 ("ARVT","Aravalli Steel & Tubes Ltd","Materials",9474,1090,395,-110,4.01,302.4421,5332,1643,3.0),
 ("GDVR","Godavari Cement Ltd","Materials",2412,277,110,38,5.98,100.3382,888,962,5.4),
 ("VRDN","Verdant Materials Plc","Materials",3354,386,154,45,3.57,169.1968,1250,1092,6.0),
 ("CHNB","Chenab Power Ltd","Utilities",8463,1989,1073,748,95.42,117.2645,2265,4985,5.6),
 ("SRYU","Suryodaya Solar Ltd","Utilities",3941,926,515,359,18.49,313.4833,1031,2045,0.9),
]
PEER_TICKERS = ["BHMT", "NLSP", "PLMT", "HRBR", "TPTI"]

# ---------------------------------------------------------------------------
# Styling -- the banking conventions taught in Module 14
# ---------------------------------------------------------------------------
FONT = "Arial"
BLUE  = Font(name=FONT, size=10, color="0000FF")               # hardcoded input
BLACK = Font(name=FONT, size=10, color="000000")               # formula
GREEN = Font(name=FONT, size=10, color="008000")               # link to another sheet
BOLD  = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=13, bold=True)
HEAD  = Font(name=FONT, size=10, bold=True, color="FFFFFF")
NOTE  = Font(name=FONT, size=9, italic=True, color="595959")
HEADFILL = PatternFill("solid", fgColor="1F3864")
SECFILL  = PatternFill("solid", fgColor="D9E2F3")
WARNFILL = PatternFill("solid", fgColor="FFF2CC")
TOPBORDER = Border(top=Side(style="thin"))

CUR  = '#,##0;(#,##0);"-"'
CUR1 = '#,##0.0;(#,##0.0);"-"'
PCT  = '0.0%;(0.0%);"-"'
DAY  = '0.0'
MULT = '0.0x'
PX   = '#,##0.00'


def put(ws, cell, value, font=BLACK, fmt=None, fill=None, align=None):
    c = ws[cell]
    c.value = value
    c.font = font
    if fmt:   c.number_format = fmt
    if fill:  c.fill = fill
    if align: c.alignment = Alignment(horizontal=align)
    return c


def section(ws, row, text, last_col):
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).fill = SECFILL
    put(ws, f"A{row}", text, BOLD, fill=SECFILL)


def col_header(ws, row, labels, start_col=2):
    for i, lab in enumerate(labels):
        put(ws, f"{get_column_letter(start_col + i)}{row}", lab, HEAD,
            fill=HEADFILL, align="center")


def readme_sheet(wb, title, lines):
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 108
    put(ws, "A1", title, TITLE)
    r = 3
    for line in lines:
        f = BOLD if line.startswith("#") else Font(name=FONT, size=10)
        put(ws, f"A{r}", line.lstrip("# "), f)
        ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return ws


LEGEND = [
    "# How to read this workbook",
    "Blue text  = a hardcoded input. Every assumption in the model is blue, and blue is the only",
    "             font colour you should ever type over.",
    "Black text = a formula. If you type over black text you have broken the model.",
    "Green text = a link to another sheet in this workbook.",
    "Yellow fill = the checks block. It must read zero.",
    "",
]


# ===========================================================================
# WORKBOOK 1 -- three-statement model
# ===========================================================================
def build_three_statement(path):
    wb = Workbook(); wb.remove(wb.active)

    # ---- Assumptions -----------------------------------------------------
    a = wb.create_sheet("Assumptions")
    a.column_dimensions["A"].width = 46
    for col in "BCDEFG": a.column_dimensions[col].width = 12
    put(a, "A1", f"{COMPANY['name']} ({COMPANY['ticker']}) - forecast assumptions", TITLE)
    put(a, "A2", "Fiscal year ends 31 March. All amounts in Rs millions unless stated.", NOTE)
    put(a, "A4", "Every forecast driver is anchored to what the business actually did in FY26.", NOTE)
    put(a, "A5", "If you change an assumption, say why in the note column.", NOTE)
    put(a, "A7", "Driver", BOLD)
    col_header(a, 7, ["FY26A", "FY27E", "FY28E", "FY29E", "FY30E", "FY31E"])

    rows = [
        # label, FY26A formula (green link or computed), FY27..FY31 inputs, format
        ("Revenue growth (%)",                    "=Model!D7",  [0.060,0.055,0.050,0.045,0.040], PCT),
        ("Gross margin (%)",                      "=Model!D11", [0.420,0.420,0.421,0.421,0.422], PCT),
        ("Operating expenses (% of revenue)",     "=Model!D14", [0.285,0.285,0.284,0.284,0.283], PCT),
        ("Depreciation (% of opening net PPE)",   "=Model!D46", [0.080,0.080,0.080,0.080,0.080], PCT),
        ("Amortisation (% of opening intangibles)","=Model!D53",[0.060,0.060,0.060,0.060,0.060], PCT),
        ("Capex (% of revenue)",                  "=Model!D47", [0.050,0.050,0.049,0.049,0.048], PCT),
        ("Days sales outstanding (DSO)",          "=Model!D28", [30.6,30.6,30.0,30.0,30.0],      DAY),
        ("Days inventory outstanding (DIO)",      "=Model!D30", [76.6,76.0,75.0,75.0,75.0],      DAY),
        ("Days payable outstanding (DPO)",        "=Model!D33", [51.1,51.0,51.0,51.0,51.0],      DAY),
        ("Other current assets (% of revenue)",   "=Model!D35", [0.0128]*5,                      PCT),
        ("Accrued liabilities (% of revenue)",    "=Model!D36", [0.0102]*5,                      PCT),
        ("Deferred revenue (% of revenue)",       "=Model!D37", [0.0128]*5,                      PCT),
        ("Interest rate on opening total debt (%)","=Model!D60",[0.100]*5,                       PCT),
        ("Interest earned on opening cash (%)",   "=Model!D61", [0.024]*5,                       PCT),
        ("Effective tax rate (%)",                "=Model!D62", [0.250]*5,                       PCT),
        ("Scheduled debt repayment (Rs m)",       223,          [223]*5,                         CUR),
        ("Dividend payout (% of net income)",     0.0,          [0.0]*5,                         PCT),
    ]
    r = 8
    for label, hist, fc, fmt in rows:
        put(a, f"A{r}", label)
        if isinstance(hist, str):
            put(a, f"B{r}", hist, GREEN, fmt)
        else:
            put(a, f"B{r}", hist, BLUE, fmt)
        for i, v in enumerate(fc):
            put(a, f"{get_column_letter(3+i)}{r}", v, BLUE, fmt)
        r += 1
    put(a, f"A{r+1}", "Blue = you may change it. Green = pulled from the Model sheet's actuals.", NOTE)
    put(a, f"A{r+2}", "Note: interest is charged on OPENING debt, not average debt. That removes the", NOTE)
    put(a, f"A{r+3}", "circular reference entirely, at the cost of a little precision. See Module 14.", NOTE)

    # ---- Model -----------------------------------------------------------
    m = wb.create_sheet("Model")
    m.column_dimensions["A"].width = 46
    for col in "BCDEFGHI": m.column_dimensions[col].width = 12
    m.freeze_panes = "B6"
    put(m, "A1", f"{COMPANY['name']} ({COMPANY['ticker']}) - three-statement model", TITLE)
    put(m, "A2", "Rs millions, fiscal years ending 31 March", NOTE)
    put(m, "A3", "FY24-FY26 are actuals (blue). FY27-FY31 are formulas, identical across every "
                 "forecast column.", NOTE)
    put(m, "A5", "Fiscal year", BOLD)
    col_header(m, 5, ["FY24A","FY25A","FY26A","FY27E","FY28E","FY29E","FY30E","FY31E"])

    H = ["B","C","D"]          # historical columns
    F = ["E","F","G","H","I"]  # forecast columns
    ALL = H + F
    AC = ["C","D","E","F","G"]  # matching Assumptions columns for FY27..FY31 is C..G

    def hist(row, key, src, fmt=CUR):
        for i, c in enumerate(H):
            put(m, f"{c}{row}", src[key][i], BLUE, fmt)

    def across(row, tmpl, cols, font=BLACK, fmt=CUR):
        """tmpl uses {c}=this column, {p}=previous column, {a}=assumptions column."""
        for i, c in enumerate(cols):
            prev = ALL[ALL.index(c) - 1]
            acol = AC[i] if cols is F else None
            put(m, f"{c}{row}", tmpl.format(c=c, p=prev, a=acol), font, fmt)

    # --- income statement
    section(m, 6, "INCOME STATEMENT", 9)
    put(m, "A7", "Revenue")
    hist(7, "revenue", HIST_IS)
    across(7, "={p}7*(1+Assumptions!{a}8)", F)
    put(m, "A8", "  Revenue growth (%)")
    across(8, "={c}7/{p}7-1", ["C","D"] + F, fmt=PCT)
    put(m, "A9", "Cost of goods sold")
    hist(9, "cogs", HIST_IS)
    across(9, "={c}7*(1-Assumptions!{a}9)", F)
    put(m, "A10", "Gross profit", BOLD)
    across(10, "={c}7-{c}9", ALL, BOLD)
    put(m, "A11", "  Gross margin (%)")
    across(11, "={c}10/{c}7", ALL, fmt=PCT)
    put(m, "A12", "Operating expenses")
    hist(12, "opex", HIST_IS)
    across(12, "={c}7*Assumptions!{a}10", F)
    put(m, "A13", "EBITDA", BOLD)
    across(13, "={c}10-{c}12", ALL, BOLD)
    put(m, "A14", "  EBITDA margin (%)")
    across(14, "={c}13/{c}7", ALL, fmt=PCT)
    put(m, "A15", "Depreciation")
    hist(15, "dep", HIST_IS)
    across(15, "={c}48", F)
    put(m, "A16", "Amortisation")
    hist(16, "amort", HIST_IS)
    across(16, "={c}54", F)
    put(m, "A17", "EBIT", BOLD)
    across(17, "={c}13-{c}15-{c}16", ALL, BOLD)
    put(m, "A18", "Interest expense")
    hist(18, "int_exp", HIST_IS)
    across(18, "={p}64*Assumptions!{a}20", F)
    put(m, "A19", "Interest income")
    hist(19, "int_inc", HIST_IS)
    across(19, "={p}82*Assumptions!{a}21", F)
    put(m, "A20", "Profit before tax", BOLD)
    across(20, "={c}17-{c}18+{c}19", ALL, BOLD)
    put(m, "A21", "Tax")
    hist(21, "tax", HIST_IS)
    across(21, "=MAX({c}20,0)*Assumptions!{a}22", F)
    put(m, "A22", "Net income", BOLD)
    across(22, "={c}20-{c}21", ALL, BOLD)
    put(m, "A23", "  Net margin (%)")
    across(23, "={c}22/{c}7", ALL, fmt=PCT)
    put(m, "A24", "Dividends paid")
    for c in H: put(m, f"{c}24", 0, BLUE, CUR)
    across(24, "=MAX({c}22,0)*Assumptions!{a}24", F)

    # --- working capital
    section(m, 26, "WORKING CAPITAL SCHEDULE", 9)
    put(m, "A27", "Accounts receivable")
    hist(27, "ar", HIST_BS)
    across(27, "=Assumptions!{a}14/365*{c}7", F)
    put(m, "A28", "  DSO (days)")
    across(28, "={c}27/{c}7*365", ALL, fmt=DAY)
    put(m, "A29", "Inventory")
    hist(29, "inv", HIST_BS)
    across(29, "=Assumptions!{a}15/365*{c}9", F)
    put(m, "A30", "  DIO (days)")
    across(30, "={c}29/{c}9*365", ALL, fmt=DAY)
    put(m, "A31", "Other current assets")
    hist(31, "oca", HIST_BS)
    across(31, "={c}7*Assumptions!{a}17", F)
    put(m, "A32", "Accounts payable")
    hist(32, "ap", HIST_BS)
    across(32, "=Assumptions!{a}16/365*{c}9", F)
    put(m, "A33", "  DPO (days)")
    across(33, "={c}32/{c}9*365", ALL, fmt=DAY)
    put(m, "A34", "Accrued liabilities")
    hist(34, "accr", HIST_BS)
    across(34, "={c}7*Assumptions!{a}18", F)
    put(m, "A35", "  Other current assets (% of revenue)")
    across(35, "={c}31/{c}7", ALL, fmt=PCT)
    put(m, "A36", "  Accrued liabilities (% of revenue)")
    across(36, "={c}34/{c}7", ALL, fmt=PCT)
    put(m, "A37", "  Deferred revenue (% of revenue)")
    across(37, "={c}38/{c}7", ALL, fmt=PCT)
    put(m, "A38", "Deferred revenue")
    hist(38, "dr", HIST_BS)
    across(38, "={c}7*Assumptions!{a}19", F)
    put(m, "A39", "Net working capital", BOLD)
    across(39, "={c}27+{c}29+{c}31-{c}32-{c}34-{c}38", ALL, BOLD)
    put(m, "A40", "Change in net working capital")
    across(40, "={c}39-{p}39", ["C","D"] + F)
    put(m, "A41", "Cash impact of working capital")
    across(41, "=-{c}40", ["C","D"] + F)

    # --- PPE
    section(m, 43, "FIXED ASSET AND INTANGIBLE SCHEDULES", 9)
    put(m, "A44", "Opening net PPE")
    for c in ["C","D"]:
        prev = ALL[ALL.index(c) - 1]
        put(m, f"{c}44", f"={prev}47", GREEN if False else BLACK, CUR)
    across(44, "={p}47", F)
    put(m, "A45", "Capex")
    for i, c in enumerate(H): put(m, f"{c}45", HIST_CF["capex"][i], BLUE, CUR)
    across(45, "={c}7*Assumptions!{a}13", F)
    put(m, "A46", "  Depreciation (% of opening PPE)")
    across(46, "={c}48/{c}44", ["C","D"] + F, fmt=PCT)
    put(m, "A47", "Closing net PPE", BOLD)
    for i, c in enumerate(H): put(m, f"{c}47", HIST_BS["ppe"][i], BLUE, CUR)
    across(47, "={c}44+{c}45-{c}48", F, BOLD)
    put(m, "A48", "Depreciation")
    for i, c in enumerate(H): put(m, f"{c}48", HIST_IS["dep"][i], BLUE, CUR)
    across(48, "={c}44*Assumptions!{a}11", F)
    put(m, "A49", "  Capex (% of revenue) [historical check]")
    across(49, "={c}45/{c}7", ALL, fmt=PCT)

    put(m, "A51", "Opening intangibles")
    for c in ["C","D"]:
        prev = ALL[ALL.index(c) - 1]
        put(m, f"{c}51", f"={prev}55", BLACK, CUR)
    across(51, "={p}55", F)
    put(m, "A52", "Amortisation")
    for i, c in enumerate(H): put(m, f"{c}52", HIST_IS["amort"][i], BLUE, CUR)
    across(52, "={c}54", F)
    put(m, "A53", "  Amortisation (% of opening intangibles)")
    across(53, "={c}52/{c}51", ["C","D"] + F, fmt=PCT)
    put(m, "A54", "Amortisation charge")
    for i, c in enumerate(H): put(m, f"{c}54", HIST_IS["amort"][i], BLUE, CUR)
    across(54, "={c}51*Assumptions!{a}12", F)
    put(m, "A55", "Closing intangibles", BOLD)
    for i, c in enumerate(H): put(m, f"{c}55", HIST_BS["gi"][i], BLUE, CUR)
    across(55, "={c}51-{c}54", F, BOLD)

    # --- debt
    section(m, 57, "DEBT SCHEDULE", 9)
    put(m, "A58", "Opening long-term debt")
    for c in ["C","D"]:
        prev = ALL[ALL.index(c) - 1]
        put(m, f"{c}58", f"={prev}61", BLACK, CUR)
    across(58, "={p}61", F)
    put(m, "A59", "Scheduled repayment")
    for i, c in enumerate(H): put(m, f"{c}59", HIST_CF["debt_repaid"][i], BLUE, CUR)
    across(59, "=MIN(Assumptions!{a}23,{c}58)", F)
    put(m, "A60", "  Interest rate on opening total debt (%)")
    across(60, "={c}18/{p}64", ["C","D"] + F, fmt=PCT)
    put(m, "A61", "Closing long-term debt", BOLD)
    for i, c in enumerate(H): put(m, f"{c}61", HIST_BS["ltd"][i], BLUE, CUR)
    across(61, "={c}58-{c}59", F, BOLD)
    put(m, "A62", "  Effective tax rate (%)")
    across(62, "={c}21/{c}20", ALL, fmt=PCT)
    put(m, "A63", "Short-term debt")
    for i, c in enumerate(H): put(m, f"{c}63", HIST_BS["std"][i], BLUE, CUR)
    across(63, "={p}63", F)
    put(m, "A64", "Total debt", BOLD)
    across(64, "={c}61+{c}63", ALL, BOLD)

    # --- cash flow (FY25 onward: FY24 has no opening balance sheet)
    section(m, 66, "CASH FLOW STATEMENT", 9)
    CF = ["C","D"] + F
    put(m, "A67", "Net income")
    across(67, "={c}22", CF)
    put(m, "A68", "Add back: depreciation and amortisation")
    across(68, "={c}15+{c}16", CF)
    put(m, "A69", "Change in working capital")
    across(69, "={c}41", CF)
    put(m, "A70", "Cash from operations", BOLD)
    across(70, "=SUM({c}67:{c}69)", CF, BOLD)
    put(m, "A71", "Capital expenditure")
    across(71, "=-{c}45", CF)
    put(m, "A72", "Cash from investing", BOLD)
    across(72, "={c}71", CF, BOLD)
    put(m, "A73", "Debt repayment")
    across(73, "=-{c}59", CF)
    put(m, "A74", "Dividends paid")
    across(74, "=-{c}24", CF)
    put(m, "A75", "Cash from financing", BOLD)
    across(75, "=SUM({c}73:{c}74)", CF, BOLD)
    put(m, "A76", "Other, including rounding")
    # History is re-derived from balance sheet actuals that are rounded to whole
    # millions, so it does not tie to the rupee. Real models put that difference
    # in a labelled line rather than nudging a number until it disappears.
    # In the forecast it is hardcoded to zero, so there is no hidden plug.
    for i, c in enumerate(["C", "D"]):
        prev = ALL[ALL.index(c) - 1]
        put(m, f"{c}76", f"={c}82-({prev}82+{c}70+{c}72+{c}75)", BLACK, CUR)
    for c in F: put(m, f"{c}76", 0, BLUE, CUR)
    put(m, "A77", "Net change in cash", BOLD)
    across(77, "={c}70+{c}72+{c}75+{c}76", CF, BOLD)
    put(m, "A78", "Opening cash")
    across(78, "={p}82", CF)
    put(m, "A79", "Closing cash", BOLD)
    across(79, "={c}78+{c}77", CF, BOLD)
    put(m, "A80", "  (memo) unlevered free cash flow")
    across(80, "={c}17*(1-{c}62)+{c}15+{c}16-{c}45+{c}41", CF)

    # --- balance sheet
    section(m, 81, "BALANCE SHEET", 9)
    bs = [("Cash", 82, "cash", "={c}79"),
          ("Accounts receivable", 83, "ar", "={c}27"),
          ("Inventory", 84, "inv", "={c}29"),
          ("Other current assets", 85, "oca", "={c}31"),
          ("Net property, plant and equipment", 86, "ppe", "={c}47"),
          ("Goodwill and intangibles", 87, "gi", "={c}55"),
          ("Other assets", 88, "oa", "={p}88")]
    for label, row, key, tmpl in bs:
        put(m, f"A{row}", label)
        for i, c in enumerate(H): put(m, f"{c}{row}", HIST_BS[key][i], BLUE, CUR)
        across(row, tmpl, F)
    put(m, "A89", "Total assets", BOLD)
    across(89, "=SUM({c}82:{c}88)", ALL, BOLD)
    for c in ALL: m[f"{c}89"].border = TOPBORDER

    bl = [("Accounts payable", 90, "ap", "={c}32"),
          ("Accrued liabilities", 91, "accr", "={c}34"),
          ("Deferred revenue", 92, "dr", "={c}38"),
          ("Short-term debt", 93, "std", "={c}63"),
          ("Long-term debt", 94, "ltd", "={c}61"),
          ("Other liabilities", 95, "ol", "={p}95")]
    for label, row, key, tmpl in bl:
        put(m, f"A{row}", label)
        for i, c in enumerate(H): put(m, f"{c}{row}", HIST_BS[key][i], BLUE, CUR)
        across(row, tmpl, F)
    put(m, "A96", "Total liabilities", BOLD)
    across(96, "=SUM({c}90:{c}95)", ALL, BOLD)
    put(m, "A97", "Share capital")
    for i, c in enumerate(H): put(m, f"{c}97", HIST_BS["sc"][i], BLUE, CUR)
    across(97, "={p}97", F)
    put(m, "A98", "Retained earnings")
    for i, c in enumerate(H): put(m, f"{c}98", HIST_BS["re"][i], BLUE, CUR)
    across(98, "={p}98+{c}22-{c}24", F)
    put(m, "A99", "Total equity", BOLD)
    across(99, "={c}97+{c}98", ALL, BOLD)
    put(m, "A100", "Total liabilities and equity", BOLD)
    across(100, "={c}96+{c}99", ALL, BOLD)
    for c in ALL: m[f"{c}100"].border = TOPBORDER

    # --- checks
    section(m, 102, "CHECKS  (every cell must read zero)", 9)
    put(m, "A103", "1. Balance sheet: total assets less total liabilities and equity")
    across(103, "={c}89-{c}100", ALL)
    put(m, "A104", "2. Cash link: balance sheet cash less cash flow closing cash (FY25 onwards)")
    across(104, "={c}82-{c}79", CF)
    put(m, "A105", "3. Debt link: balance sheet debt less debt schedule total debt")
    across(105, "={c}93+{c}94-{c}64", ALL)
    for row in (103, 104, 105):
        for c in ALL: m[f"{c}{row}"].fill = WARNFILL
    put(m, "A107", "MODEL OK?", BOLD)
    put(m, "B107", '=IF(SUM(ABS(B103:I105))<0.5,"OK","ERROR - see checks above")', BOLD,
        fill=WARNFILL)
    m["B107"].value = '=IF(SUMPRODUCT(ABS(B103:I105))<0.5,"OK","ERROR - see checks above")'
    put(m, "A109", "Why check 1 can be trusted: cash is NOT plugged in this model. It is built up "
                   "from the cash flow", NOTE)
    put(m, "A110", "statement, and the balance sheet is built from the schedules. If the two "
                   "disagree, something is genuinely wrong.", NOTE)

    readme_sheet(wb, f"{COMPANY['name']} - three-statement model", LEGEND + [
        "# What this is",
        "A complete, balancing three-statement forecast model. Use it as the target to compare your",
        "own build against - not as something to copy. Build yours first, then open this one.",
        "",
        "# Sheets",
        "Assumptions   every forecast driver, with the FY26 actual beside it for comparison",
        "Model         income statement, working capital, fixed assets, debt, cash flow, balance sheet, checks",
        "",
        "# Where the numbers come from",
        "FY24-FY26 actuals are pulled from the sqlcamp database (company_id = 1). The query that",
        "produced each block is quoted in scripts/build_reference_models.py.",
        "",
        "# Things worth noticing",
        "1. Cash is not a plug. It is built from the cash flow statement, and the balance sheet is",
        "   built independently from the schedules. That is why the balance check means something.",
        "2. Interest is charged on OPENING debt, not average debt. This removes the circular",
        "   reference completely. The banking alternative - average debt plus iterative calculation -",
        "   is more precise and more fragile. Module 14 covers both.",
        "3. Every forecast column (E to I) carries an identical formula. Click along row 7 and check.",
        "4. Historical columns are blue because they are actuals, not calculations.",
        "5. FY24 has no cash flow statement because there is no FY23 balance sheet to open from.",
        "",
        "# Try this",
        "Change the FY27 revenue growth on the Assumptions sheet from 6.0% to 20%. Watch every",
        "statement move and the checks stay at zero. That is what a linked model does. Then change",
        "DSO from 30.6 to 60 days and watch cash fall even though profit is unchanged - that is",
        "working capital, and it is the single most common thing juniors get wrong.",
    ])
    wb.save(path)
    return path


# ===========================================================================
# WORKBOOK 2 -- DCF and comparable companies
# ===========================================================================
# Unlevered free cash flow drivers, taken from workbook 1 (Model sheet,
# FY27E-FY31E). They are hardcoded here because linking between workbooks is
# fragile; the source cell is named beside each row.
FCF_IN = {
    "ebitda": [1317.5, 1390.0, 1481.1, 1547.8, 1633.2],
    "ebit":   [998.1, 1056.9, 1133.4, 1185.4, 1255.4],
    "da":     [319.4, 333.0, 347.7, 362.4, 377.8],
    "capex":  [488.0, 514.8, 529.7, 553.6, 564.0],
    "wc":     [-61.1, -53.1, -22.7, -53.5, -49.0],   # cash impact, negative = outflow
}
NET_DEBT_FY26 = 2412 - 3661          # short + long term debt less cash, Rs m


def build_dcf(path):
    wb = Workbook(); wb.remove(wb.active)

    # ---- Assumptions -----------------------------------------------------
    a = wb.create_sheet("Assumptions")
    a.column_dimensions["A"].width = 46
    a.column_dimensions["B"].width = 14
    a.column_dimensions["C"].width = 62
    put(a, "A1", f"{COMPANY['name']} ({COMPANY['ticker']}) - valuation assumptions", TITLE)
    put(a, "A3", "Input", BOLD); put(a, "B3", "Value", BOLD); put(a, "C3", "Where it comes from", BOLD)
    ass = [
        ("Risk-free rate",              0.070, PCT, "10-year government bond yield"),
        ("Equity risk premium",         0.060, PCT, "Long-run excess return of equities over bonds"),
        ("Levered beta",                1.10,  '0.00', "Peer average, re-levered to this capital structure"),
        ("Pre-tax cost of debt",        0.100, PCT, "Interest expense over average debt, Model!D60"),
        ("Tax rate",                    0.250, PCT, "Effective rate, Model!D62"),
        ("Market capitalisation (Rs m)", round(COMPANY['shares_m']*COMPANY['price']), CUR,
         f"{COMPANY['shares_m']}m shares x Rs {COMPANY['price']:.2f} at {COMPANY['price_date']}"),
        ("Total debt (Rs m)",           2412,  CUR, "Balance sheet at 31-Mar-2026"),
        ("Cash (Rs m)",                 3661,  CUR, "Balance sheet at 31-Mar-2026"),
        ("Terminal growth rate",        0.040, PCT, "At or below long-run nominal GDP growth"),
        ("Exit EBITDA multiple",        9.0,   MULT, "Peer median, see Comps sheet"),
        ("Diluted shares (m)",          COMPANY['shares_m'], CUR1, "dim_company.shares_out_m"),
        ("Current share price (Rs)",    COMPANY['price'], PX, f"Close on {COMPANY['price_date']}"),
    ]
    r = 4
    for label, val, fmt, src in ass:
        put(a, f"A{r}", label); put(a, f"B{r}", val, BLUE, fmt); put(a, f"C{r}", src, NOTE); r += 1

    put(a, "A17", "WACC BUILD", BOLD, fill=SECFILL)
    put(a, "A18", "Cost of equity");        put(a, "B18", "=B4+B6*B5", BLACK, PCT)
    put(a, "C18", "risk-free + beta x equity risk premium", NOTE)
    put(a, "A19", "After-tax cost of debt");put(a, "B19", "=B7*(1-B8)", BLACK, PCT)
    put(a, "A20", "Net debt");              put(a, "B20", "=B10-B11", BLACK, CUR)
    put(a, "A21", "Enterprise value (market)"); put(a, "B21", "=B9+B20", BLACK, CUR)
    put(a, "A22", "Equity weight");         put(a, "B22", "=B9/(B9+B10)", BLACK, PCT)
    put(a, "C22", "market values, not book values", NOTE)
    put(a, "A23", "Debt weight");           put(a, "B23", "=B10/(B9+B10)", BLACK, PCT)
    put(a, "A24", "WACC", BOLD);            put(a, "B24", "=B22*B18+B23*B19", BOLD, PCT)
    a["B24"].fill = WARNFILL
    put(a, "A26", "Note: this company holds more cash than debt, so net debt is negative and the",
        NOTE)
    put(a, "A27", "equity bridge ADDS the net cash back. Getting that sign wrong is a classic error.",
        NOTE)

    # ---- DCF -------------------------------------------------------------
    d = wb.create_sheet("DCF")
    d.column_dimensions["A"].width = 46
    for col in "BCDEFG": d.column_dimensions[col].width = 13
    put(d, "A1", f"{COMPANY['name']} - discounted cash flow", TITLE)
    put(d, "A2", "Rs millions. Unlevered free cash flow, mid-year discounting.", NOTE)
    put(d, "A4", "Fiscal year", BOLD)
    col_header(d, 4, ["FY27E", "FY28E", "FY29E", "FY30E", "FY31E"])
    C = ["B", "C", "D", "E", "F"]

    section(d, 5, "UNLEVERED FREE CASH FLOW", 6)
    def inrow(row, label, key, note=""):
        put(d, f"A{row}", label)
        for i, c in enumerate(C): put(d, f"{c}{row}", FCF_IN[key][i], BLUE, CUR)
        if note: put(d, f"G{row}", note, NOTE)
    inrow(6, "EBITDA", "ebitda", "from Model!E13:I13")
    inrow(7, "EBIT", "ebit", "from Model!E17:I17")
    put(d, "A8", "Tax on EBIT")
    for c in C: put(d, f"{c}8", f"=-{c}7*Assumptions!$B$8", BLACK, CUR)
    put(d, "A9", "NOPAT", BOLD)
    for c in C: put(d, f"{c}9", f"={c}7+{c}8", BOLD, CUR)
    inrow(10, "Add: depreciation and amortisation", "da", "from Model!E15:I15 + E16:I16")
    inrow(11, "Less: capital expenditure", "capex", "from Model!E45:I45")
    put(d, "A12", "Capex (shown as an outflow)")
    for c in C: put(d, f"{c}12", f"=-{c}11", BLACK, CUR)
    inrow(13, "Change in working capital (cash impact)", "wc", "from Model!E41:I41")
    put(d, "A14", "Unlevered free cash flow", BOLD)
    for c in C: put(d, f"{c}14", f"={c}9+{c}10+{c}12+{c}13", BOLD, CUR)
    for c in C: d[f"{c}14"].border = TOPBORDER

    section(d, 16, "DISCOUNTING", 6)
    put(d, "A17", "Discount period (mid-year)")
    for i, c in enumerate(C): put(d, f"{c}17", 0.5 + i, BLUE, '0.0')
    put(d, "A18", "Discount factor")
    for c in C: put(d, f"{c}18", f"=1/(1+Assumptions!$B$24)^{c}17", BLACK, '0.000')
    put(d, "A19", "Present value of free cash flow")
    for c in C: put(d, f"{c}19", f"={c}14*{c}18", BLACK, CUR)
    put(d, "A20", "Sum of present values, forecast period", BOLD)
    put(d, "B20", "=SUM(B19:F19)", BOLD, CUR)

    section(d, 22, "TERMINAL VALUE", 6)
    put(d, "A23", "Terminal discount period")
    put(d, "B23", "=F17+0.5", BLACK, '0.0')
    put(d, "C23", "end of the final year, not mid-year", NOTE)
    put(d, "A24", "Method 1: Gordon growth", BOLD)
    put(d, "A25", "  Terminal value")
    put(d, "B25", "=$B$33*(1+Assumptions!$B$12)/(Assumptions!$B$24-Assumptions!$B$12)", BLACK, CUR)
    put(d, "A26", "  Present value of terminal value")
    put(d, "B26", "=B25/(1+Assumptions!$B$24)^B23", BLACK, CUR)
    put(d, "A27", "  Implied exit EBITDA multiple")
    put(d, "B27", "=B25/F6", BLACK, MULT)
    put(d, "A28", "Method 2: Exit multiple", BOLD)
    put(d, "A29", "  Terminal value")
    put(d, "B29", "=F6*Assumptions!$B$13", BLACK, CUR)
    put(d, "A30", "  Present value of terminal value")
    put(d, "B30", "=B29/(1+Assumptions!$B$24)^B23", BLACK, CUR)
    put(d, "A31", "  Implied terminal growth rate")
    put(d, "B31", "=(Assumptions!$B$24*B29-F14)/(B29+F14)", BLACK, PCT)
    put(d, "A32", "  Normalised terminal capex (x depreciation)")
    put(d, "B32", 1.0, BLUE, MULT)
    put(d, "C32", "in perpetuity a business only needs to replace what wears out", NOTE)
    put(d, "A33", "  Normalised terminal free cash flow", BOLD)
    put(d, "B33", "=F9+F10-F10*$B$32+F13", BOLD, CUR)
    put(d, "C33", "NOPAT + D&A - normalised capex + working capital", NOTE)
    put(d, "A34", "If those two implied numbers disagree wildly, one of your assumptions is wrong.",
        NOTE)
    put(d, "A47", "Note the terminal normalisation above. FY31 capex is 1.5x depreciation because the", NOTE)
    put(d, "A48", "company is still expanding. Growing that forever at 4% would be incoherent, so the", NOTE)
    put(d, "A49", "terminal year steps capex down to maintenance level. Forgetting this understates", NOTE)
    put(d, "A50", "value in every capital-hungry business, and it is one of the commonest DCF errors.", NOTE)

    section(d, 40, "VALUATION  (Gordon growth method)", 6)
    put(d, "A41", "Present value of forecast cash flows"); put(d, "B41", "=B20", BLACK, CUR)
    put(d, "A42", "Present value of terminal value");      put(d, "B42", "=B26", BLACK, CUR)
    put(d, "A43", "Enterprise value", BOLD);               put(d, "B43", "=B41+B42", BOLD, CUR)
    put(d, "A44", "  Terminal value as % of enterprise value")
    put(d, "B44", "=B42/B43", BLACK, PCT); d["B44"].fill = WARNFILL
    put(d, "C44", "the honesty check - how much of your answer is a guess about year 6 onwards?", NOTE)
    put(d, "A45", "  Implied EV / FY26 EBITDA")
    put(d, "B45", "=B43/1243", BLACK, MULT)
    put(d, "C45", "compare this with the peer median on the Comps sheet", NOTE)
    put(d, "A46", "Less: net debt (add back net cash)")
    put(d, "B46", "=-Assumptions!B20", BLACK, CUR)
    put(d, "A47", "Equity value", BOLD);                   put(d, "B47", "=B43+B46", BOLD, CUR)
    put(d, "A48", "Diluted shares (m)");                   put(d, "B48", "=Assumptions!B14", GREEN, CUR1)
    put(d, "A49", "Value per share (Rs)", BOLD);           put(d, "B49", "=B47/B48", BOLD, PX)
    d["B49"].fill = WARNFILL
    put(d, "A50", "Current share price (Rs)");             put(d, "B50", "=Assumptions!B15", GREEN, PX)
    put(d, "A51", "Upside / (downside)", BOLD);            put(d, "B51", "=B49/B50-1", BOLD, PCT)

    section(d, 53, "VALUATION  (exit multiple method)", 6)
    put(d, "A54", "Enterprise value");           put(d, "B54", "=B20+B30", BLACK, CUR)
    put(d, "A55", "Equity value");               put(d, "B55", "=B54-Assumptions!B20", BLACK, CUR)
    put(d, "A56", "Value per share (Rs)", BOLD); put(d, "B56", "=B55/B48", BOLD, PX)

    section(d, 58, "REVERSE DCF  - what does today's price already assume?", 6)
    put(d, "A59", "This is the most honest question in valuation. Instead of asking what the share", NOTE)
    put(d, "A60", "is worth, it asks what you would have to believe to pay today's price.", NOTE)
    put(d, "A61", "Market equity value");        put(d, "B61", "=Assumptions!B9", GREEN, CUR)
    put(d, "A62", "Market enterprise value");    put(d, "B62", "=B61+Assumptions!B20", BLACK, CUR)
    put(d, "A63", "Terminal value the market implies (undiscounted)")
    put(d, "B63", "=(B62-B20)*(1+Assumptions!$B$24)^B23", BLACK, CUR)
    put(d, "A64", "Implied terminal growth rate", BOLD)
    put(d, "B64", "=(B63*Assumptions!$B$24-$B$33)/(B63+$B$33)", BOLD, PCT)
    d["B64"].fill = WARNFILL
    put(d, "A65", "Implied exit EBITDA multiple", BOLD)
    put(d, "B65", "=B63/F6", BOLD, MULT); d["B65"].fill = WARNFILL
    put(d, "A66", "Now judge those two numbers, not the share price. If the market implies terminal", NOTE)
    put(d, "A67", "growth above nominal GDP forever, or an exit multiple far above where the peers", NOTE)
    put(d, "A68", "trade today, the price is carrying an assumption somebody should defend.", NOTE)

    # ---- Sensitivity -----------------------------------------------------
    s = wb.create_sheet("Sensitivity")
    s.column_dimensions["A"].width = 24
    for col in "BCDEFG": s.column_dimensions[col].width = 12
    put(s, "A1", "Sensitivity - value per share (Rs)", TITLE)
    put(s, "A2", "Every cell recomputes the whole valuation. Nothing here is pasted.", NOTE)

    put(s, "A4", "WACC (down) vs terminal growth (across)", BOLD, fill=SECFILL)
    growths = [0.030, 0.035, 0.040, 0.045, 0.050]
    for j, g in enumerate(growths):
        put(s, f"{get_column_letter(2+j)}5", g, BLUE, PCT, align="center")
    put(s, "A5", "WACC", HEAD, fill=HEADFILL, align="center")
    waccs = [0.080, 0.090, 0.100, 0.110, 0.120, 0.130, 0.140]
    for i, w in enumerate(waccs):
        row = 6 + i
        put(s, f"A{row}", w, BLUE, PCT, align="center")
        for j in range(len(growths)):
            col = get_column_letter(2 + j)
            f = (f"=(SUMPRODUCT(DCF!$B$14:$F$14,(1+$A{row})^-DCF!$B$17:$F$17)"
                 f"+DCF!$F$14*(1+{col}$5)/($A{row}-{col}$5)/(1+$A{row})^DCF!$B$23"
                 f"-Assumptions!$B$20)/Assumptions!$B$14")
            put(s, f"{col}{row}", f, BLACK, PX)

    put(s, "A15", "WACC (down) vs exit EBITDA multiple (across)", BOLD, fill=SECFILL)
    mults = [7.0, 8.0, 9.0, 10.0, 11.0]
    for j, mu in enumerate(mults):
        put(s, f"{get_column_letter(2+j)}16", mu, BLUE, MULT, align="center")
    put(s, "A16", "WACC", HEAD, fill=HEADFILL, align="center")
    for i, w in enumerate(waccs):
        row = 17 + i
        put(s, f"A{row}", w, BLUE, PCT, align="center")
        for j in range(len(mults)):
            col = get_column_letter(2 + j)
            f = (f"=(SUMPRODUCT(DCF!$B$14:$F$14,(1+$A{row})^-DCF!$B$17:$F$17)"
                 f"+DCF!$F$6*{col}$16/(1+$A{row})^DCF!$B$23"
                 f"-Assumptions!$B$20)/Assumptions!$B$14")
            put(s, f"{col}{row}", f, BLACK, PX)
    put(s, "A26", "Read these as a range, not a point. A DCF gives you a region of plausible", NOTE)
    put(s, "A27", "values; anyone quoting a single number to two decimal places is overselling.", NOTE)

    # ---- Comps -----------------------------------------------------------
    c = wb.create_sheet("Comps")
    widths = {"A": 11, "B": 30, "C": 24}
    for k, v in widths.items(): c.column_dimensions[k].width = v
    for col in "DEFGHIJKLMNOPQ": c.column_dimensions[col].width = 11
    put(c, "A1", "Comparable companies - LTM to 31 March 2026", TITLE)
    put(c, "A2", "Rs millions except per-share and multiples. Blue = input, black = calculated.", NOTE)
    heads = ["Ticker","Company","Sector","LTM revenue","LTM EBITDA","LTM EBIT","LTM net income",
             "Shares (m)","Price (Rs)","Gross debt","Cash","Rev growth %",
             "Market cap","Net debt","Enterprise value","EV/Revenue","EV/EBITDA","EV/EBIT","P/E",
             "EBITDA margin"]
    for j, h in enumerate(heads):
        put(c, f"{get_column_letter(1+j)}4", h, HEAD, fill=HEADFILL, align="center")
    r = 5
    for row in COMPS:
        tick, name, sec, rev, ebitda, ebit, ni, sh, px, debt, cash, growth = row
        is_target = tick == COMPANY["ticker"]
        put(c, f"A{r}", tick, BOLD if is_target else BLUE)
        put(c, f"B{r}", name, BOLD if is_target else BLUE)
        put(c, f"C{r}", sec, BLUE)
        for j, v in enumerate([rev, ebitda, ebit, ni]):
            put(c, f"{get_column_letter(4+j)}{r}", v, BLUE, CUR)
        put(c, f"H{r}", sh, BLUE, CUR1)
        put(c, f"I{r}", px, BLUE, PX)
        put(c, f"J{r}", debt, BLUE, CUR)
        put(c, f"K{r}", cash, BLUE, CUR)
        put(c, f"L{r}", growth / 100, BLUE, PCT)
        put(c, f"M{r}", f"=H{r}*I{r}", BLACK, CUR)
        put(c, f"N{r}", f"=J{r}-K{r}", BLACK, CUR)
        put(c, f"O{r}", f"=M{r}+N{r}", BLACK, CUR)
        put(c, f"P{r}", f"=IF(D{r}>0,O{r}/D{r},\"nm\")", BLACK, MULT)
        put(c, f"Q{r}", f"=IF(E{r}>0,O{r}/E{r},\"nm\")", BLACK, MULT)
        put(c, f"R{r}", f"=IF(F{r}>0,O{r}/F{r},\"nm\")", BLACK, MULT)
        put(c, f"S{r}", f"=IF(G{r}>0,M{r}/G{r},\"nm\")", BLACK, MULT)
        put(c, f"T{r}", f"=E{r}/D{r}", BLACK, PCT)
        if is_target:
            for col in "ABCDEFGHIJKLMNOPQRST":
                c[f"{col}{r}"].fill = WARNFILL
        r += 1
    last = r - 1
    put(c, f"A{r+1}", "Note: a loss-making company shows \"nm\" for P/E. That is not a formatting", NOTE)
    put(c, f"A{r+2}", "problem - it is the reason EV/EBITDA exists. See ARVT.", NOTE)

    pr = r + 4
    put(c, f"A{pr}", "SELECTED PEER SET", BOLD, fill=SECFILL)
    put(c, f"C{pr}", "Chosen for similar sector, size and margin structure. Justify every "
                     "inclusion AND every exclusion.", NOTE)
    for j, h in enumerate(["Ticker","Company","","LTM revenue","LTM EBITDA","","","","","","","Rev growth %",
                           "Market cap","","Enterprise value","EV/Revenue","EV/EBITDA","EV/EBIT","P/E","EBITDA margin"]):
        if h: put(c, f"{get_column_letter(1+j)}{pr+1}", h, HEAD, fill=HEADFILL, align="center")
    pstart = pr + 2
    for i, tick in enumerate(PEER_TICKERS):
        rr = pstart + i
        put(c, f"A{rr}", tick, BLUE)
        for col in ["B","D","E","L","M","O","P","Q","R","S","T"]:
            put(c, f"{col}{rr}",
                f"=INDEX({col}$5:{col}${last},MATCH($A{rr},$A$5:$A${last},0))", BLACK,
                CUR if col in ("D","E","M","O") else (PCT if col in ("L","T") else MULT))
        c[f"B{rr}"].number_format = "General"
    pend = pstart + len(PEER_TICKERS) - 1
    stats = [("Median", "MEDIAN"), ("Mean", "AVERAGE"), ("Minimum", "MIN"), ("Maximum", "MAX")]
    sr = pend + 1
    for name, fn in stats:
        put(c, f"A{sr}", name, BOLD)
        for col in ["P","Q","R","S"]:
            put(c, f"{col}{sr}", f"={fn}({col}{pstart}:{col}{pend})", BOLD, MULT)
        sr += 1
    put(c, f"A{sr}", "Lower quartile", BOLD)
    for col in ["P","Q","R","S"]:
        put(c, f"{col}{sr}", f"=QUARTILE({col}{pstart}:{col}{pend},1)", BOLD, MULT)
    put(c, f"A{sr+1}", "Upper quartile", BOLD)
    for col in ["P","Q","R","S"]:
        put(c, f"{col}{sr+1}", f"=QUARTILE({col}{pstart}:{col}{pend},3)", BOLD, MULT)
    med_row = pend + 1

    ir = sr + 3
    put(c, f"A{ir}", f"IMPLIED VALUE OF {COMPANY['ticker']} AT THE PEER MEDIAN", BOLD, fill=SECFILL)
    kv = [i for i, row in enumerate(COMPS) if row[0] == COMPANY["ticker"]][0] + 5
    put(c, f"A{ir+1}", "Peer median EV/EBITDA");      put(c, f"D{ir+1}", f"=Q{med_row}", GREEN, MULT)
    put(c, f"A{ir+2}", f"{COMPANY['ticker']} LTM EBITDA"); put(c, f"D{ir+2}", f"=E{kv}", GREEN, CUR)
    put(c, f"A{ir+3}", "Implied enterprise value");   put(c, f"D{ir+3}", f"=D{ir+1}*D{ir+2}", BLACK, CUR)
    put(c, f"A{ir+4}", "Less: net debt");             put(c, f"D{ir+4}", f"=-N{kv}", BLACK, CUR)
    put(c, f"A{ir+5}", "Implied equity value");       put(c, f"D{ir+5}", f"=D{ir+3}+D{ir+4}", BOLD, CUR)
    put(c, f"A{ir+6}", "Implied value per share (Rs)", BOLD)
    put(c, f"D{ir+6}", f"=D{ir+5}/H{kv}", BOLD, PX); c[f"D{ir+6}"].fill = WARNFILL
    put(c, f"A{ir+7}", "Current share price (Rs)");   put(c, f"D{ir+7}", f"=I{kv}", GREEN, PX)
    put(c, f"A{ir+8}", "Upside / (downside)", BOLD);  put(c, f"D{ir+8}", f"=D{ir+6}/D{ir+7}-1", BOLD, PCT)

    readme_sheet(wb, f"{COMPANY['name']} - DCF and comparable companies", LEGEND + [
        "# What this is",
        "The valuation half of the reference set: a discounted cash flow, two sensitivity grids and a",
        "comparable companies table. It follows on from workbook 01.",
        "",
        "# Sheets",
        "Assumptions   WACC build and the valuation inputs, each with its source",
        "DCF           unlevered free cash flow, discounting, both terminal value methods, the equity bridge",
        "Sensitivity   value per share across WACC x terminal growth, and WACC x exit multiple",
        "Comps         all 20 companies, a selected peer set, medians and quartiles, implied value",
        "",
        "# Things worth noticing",
        "1. The free cash flow lines are hardcoded (blue) because they come from workbook 01. Linking",
        "   between separate Excel files breaks the moment someone renames a folder. Copy the values,",
        "   and write down which cells they came from - the note beside each row does exactly that.",
        "2. Free cash flow is UNLEVERED: there is no interest anywhere in the build. You are valuing",
        "   the business, then deciding separately who funds it. That is what the equity bridge is for.",
        "3. This company has more cash than debt. Net debt is negative, so the bridge ADDS it back.",
        "   Check the sign every single time.",
        "4. Cell DCF!B38 shows terminal value as a share of enterprise value. Look at it before you",
        "   quote a target price to anybody.",
        "5. Both terminal value methods are shown, each with the other's implied assumption. If the",
        "   Gordon method implies a 15x exit multiple while your peers trade at 9x, your growth rate",
        "   is fantasy.",
        "6. Every sensitivity cell recalculates the entire valuation from first principles. Change the",
        "   WACC on the Assumptions sheet and the grids do not move - they are deliberately",
        "   independent, so they show the whole landscape rather than one point on it.",
        "",
        "# Try this",
        "Set the terminal growth rate to 6% and watch the value per share. Then ask yourself whether",
        "you believe a retailer grows faster than the economy forever. That is the question a DCF",
        "actually asks you.",
    ])
    wb.save(path)
    return path


# ===========================================================================
# WORKBOOK 3 -- leveraged buyout
# ===========================================================================
def build_lbo(path):
    wb = Workbook(); wb.remove(wb.active)
    C = ["B", "C", "D", "E", "F"]          # FY27E .. FY31E

    # ---- Assumptions -----------------------------------------------------
    a = wb.create_sheet("Assumptions")
    a.column_dimensions["A"].width = 46
    a.column_dimensions["B"].width = 14
    a.column_dimensions["C"].width = 60
    put(a, "A1", f"{COMPANY['name']} ({COMPANY['ticker']}) - LBO assumptions", TITLE)
    put(a, "A2", "Rs millions. Assumed close 31-Mar-2026, five-year hold to 31-Mar-2031.", NOTE)
    put(a, "A4", "Input", BOLD); put(a, "B4", "Value", BOLD); put(a, "C4", "Note", BOLD)
    ass = [
        ("LTM EBITDA at entry",        1243,  CUR,  "FY26 actual"),
        ("Entry EV / EBITDA multiple", 11.0,  MULT, "What you are willing to pay. The premium falls out of this."),
        ("Entry leverage (x EBITDA)",  4.0,   MULT, "How much debt the lenders will provide"),
        ("Transaction fees (% of EV)", 0.020, PCT,  "Advisory, financing and legal"),
        ("Interest rate on term loan", 0.110, PCT,  "Floating rate plus margin"),
        ("Mandatory amortisation (% of original)", 0.010, PCT, "Contractual repayment each year"),
        ("Cash sweep (% of surplus)",  1.00,  PCT,  "Share of free cash flow that must repay debt"),
        ("Tax rate",                   0.250, PCT,  "Effective rate"),
        ("Exit EV / EBITDA multiple",  11.0,  MULT, "Assume you sell at what you paid. Anything better is luck."),
        ("Hold period (years)",        5,     '0',  ""),
        ("Existing debt refinanced",   2412,  CUR,  "Balance sheet at 31-Mar-2026"),
        ("Cash on the balance sheet",  3661,  CUR,  "Balance sheet at 31-Mar-2026, used to fund the deal"),
        ("Shares outstanding (m)",     COMPANY['shares_m'], CUR1, "dim_company.shares_out_m"),
        ("Share price before the deal (Rs)", COMPANY['price'], PX, f"Close on {COMPANY['price_date']}"),
    ]
    r = 5
    for label, val, fmt, note in ass:
        put(a, f"A{r}", label); put(a, f"B{r}", val, BLUE, fmt); put(a, f"C{r}", note, NOTE); r += 1

    put(a, "A20", "DERIVED FROM THE ABOVE", BOLD, fill=SECFILL)
    put(a, "A21", "Entry enterprise value");   put(a, "B21", "=B5*B6", BLACK, CUR)
    put(a, "A22", "New term loan");            put(a, "B22", "=B5*B7", BLACK, CUR)
    put(a, "A23", "Transaction fees");         put(a, "B23", "=B21*B8", BLACK, CUR)
    put(a, "A24", "Equity purchase price");    put(a, "B24", "=B21+B16-B15", BLACK, CUR)
    put(a, "C24", "enterprise value plus cash acquired less debt assumed", NOTE)
    put(a, "A25", "Implied offer per share (Rs)", BOLD); put(a, "B25", "=B24/B17", BOLD, PX)
    put(a, "A26", "Implied premium to the undisturbed price", BOLD)
    put(a, "B26", "=B25/B18-1", BOLD, PCT); a["B26"].fill = WARNFILL
    put(a, "A27", "The premium is an OUTPUT, not an input. Raise the entry multiple and you can", NOTE)
    put(a, "A28", "offer more; the returns then have to survive it. That trade-off is the whole job.", NOTE)

    # ---- Sources and uses -----------------------------------------------
    su = wb.create_sheet("Sources and Uses")
    su.column_dimensions["A"].width = 44; su.column_dimensions["B"].width = 14
    su.column_dimensions["C"].width = 12; su.column_dimensions["D"].width = 44
    put(su, "A1", "Sources and uses of funds", TITLE)
    put(su, "A3", "USES", BOLD, fill=SECFILL); put(su, "B3", "Rs m", BOLD, fill=SECFILL)
    put(su, "C3", "% of total", BOLD, fill=SECFILL)
    put(su, "A4", "Purchase of equity");        put(su, "B4", "=Assumptions!B24", GREEN, CUR)
    put(su, "A5", "Refinance existing debt");   put(su, "B5", "=Assumptions!B15", GREEN, CUR)
    put(su, "A6", "Transaction fees");          put(su, "B6", "=Assumptions!B23", GREEN, CUR)
    put(su, "A7", "Total uses", BOLD);          put(su, "B7", "=SUM(B4:B6)", BOLD, CUR)
    for r_ in range(4, 8): put(su, f"C{r_}", f"=B{r_}/$B$7", BLACK, PCT)
    su["B7"].border = TOPBORDER

    put(su, "A9", "SOURCES", BOLD, fill=SECFILL); put(su, "B9", "Rs m", BOLD, fill=SECFILL)
    put(su, "C9", "% of total", BOLD, fill=SECFILL); put(su, "D9", "x EBITDA", BOLD, fill=SECFILL)
    put(su, "A10", "New term loan B");          put(su, "B10", "=Assumptions!B22", GREEN, CUR)
    put(su, "D10", "=B10/Assumptions!B5", BLACK, MULT)
    put(su, "A11", "Cash on the balance sheet");put(su, "B11", "=Assumptions!B16", GREEN, CUR)
    put(su, "A12", "Sponsor equity", BOLD);     put(su, "B12", "=B7-B10-B11", BOLD, CUR)
    put(su, "A13", "Total sources", BOLD);      put(su, "B13", "=SUM(B10:B12)", BOLD, CUR)
    for r_ in range(10, 14): put(su, f"C{r_}", f"=B{r_}/$B$13", BLACK, PCT)
    su["B13"].border = TOPBORDER
    put(su, "A15", "CHECK: sources less uses", BOLD)
    put(su, "B15", "=B13-B7", BOLD, CUR); su["B15"].fill = WARNFILL
    put(su, "D15", "must be zero, always, before anything else is believed", NOTE)
    put(su, "A17", "Equity as % of total capitalisation")
    put(su, "B17", "=B12/(B10+B12)", BLACK, PCT)
    put(su, "A18", "Opening net debt / EBITDA")
    put(su, "B18", "=B10/Assumptions!B5", BLACK, MULT)

    # ---- Model -----------------------------------------------------------
    m = wb.create_sheet("Model")
    m.column_dimensions["A"].width = 46
    for col in "BCDEF": m.column_dimensions[col].width = 13
    put(m, "A1", f"{COMPANY['name']} - LBO operating and debt model", TITLE)
    put(m, "A2", "Rs millions. Operating forecast carried over from the three-statement model.", NOTE)
    put(m, "A4", "Fiscal year", BOLD)
    col_header(m, 4, ["FY27E", "FY28E", "FY29E", "FY30E", "FY31E"])

    section(m, 5, "OPERATING FORECAST", 6)
    def inrow(row, label, vals, note=""):
        put(m, f"A{row}", label)
        for i, c in enumerate(C): put(m, f"{c}{row}", vals[i], BLUE, CUR)
        if note: put(m, f"G{row}", note, NOTE)
    inrow(6, "EBITDA", FCF_IN["ebitda"], "from workbook 01, Model!E13:I13")
    inrow(7, "Depreciation and amortisation", FCF_IN["da"])
    inrow(8, "Capital expenditure", FCF_IN["capex"])
    inrow(9, "Change in working capital (cash impact)", FCF_IN["wc"])
    put(m, "A10", "EBIT", BOLD)
    for c in C: put(m, f"{c}10", f"={c}6-{c}7", BOLD, CUR)

    section(m, 12, "DEBT SCHEDULE AND CASH SWEEP", 6)
    put(m, "A13", "Opening term loan")
    put(m, "B13", "='Sources and Uses'!B10", GREEN, CUR)
    for c in C[1:]:
        p_ = C[C.index(c) - 1]
        put(m, f"{c}13", f"={p_}20", BLACK, CUR)
    put(m, "A14", "Interest expense")
    for c in C: put(m, f"{c}14", f"={c}13*Assumptions!$B$9", BLACK, CUR)
    put(m, "A15", "Profit before tax")
    for c in C: put(m, f"{c}15", f"={c}10-{c}14", BLACK, CUR)
    put(m, "A16", "Tax")
    for c in C: put(m, f"{c}16", f"=MAX({c}15,0)*Assumptions!$B$12", BLACK, CUR)
    put(m, "A17", "Net income", BOLD)
    for c in C: put(m, f"{c}17", f"={c}15-{c}16", BOLD, CUR)
    put(m, "A18", "Free cash flow before debt service", BOLD)
    for c in C: put(m, f"{c}18", f"={c}17+{c}7-{c}8+{c}9", BOLD, CUR)
    put(m, "A19", "Mandatory amortisation")
    for c in C:
        put(m, f"{c}19", f"=-MIN(Assumptions!$B$10*'Sources and Uses'!$B$10,{c}13)", BLACK, CUR)
    put(m, "A20", "Cash sweep")
    for c in C:
        put(m, f"{c}20", f"=-MIN(MAX({c}18+{c}19,0)*Assumptions!$B$11,{c}13+{c}19)", BLACK, CUR)
    put(m, "A21", "Closing term loan", BOLD)
    for c in C: put(m, f"{c}21", f"={c}13+{c}19+{c}20", BOLD, CUR)
    # note: rows 19/20 are negative, row 21 adds them
    put(m, "A22", "Cash retained in the business")
    put(m, "B22", "=MAX(B18+B19+B20,0)", BLACK, CUR)
    for c in C[1:]:
        p_ = C[C.index(c) - 1]
        put(m, f"{c}22", f"={p_}22+MAX({c}18+{c}19+{c}20,0)", BLACK, CUR)
    put(m, "A23", "Net debt", BOLD)
    for c in C: put(m, f"{c}23", f"={c}21-{c}22", BOLD, CUR)
    put(m, "A24", "  Net debt / EBITDA")
    for c in C: put(m, f"{c}24", f"={c}23/{c}6", BLACK, MULT)
    put(m, "A25", "  EBITDA / interest (interest cover)")
    for c in C: put(m, f"{c}25", f"={c}6/{c}14", BLACK, MULT)
    put(m, "A26", "  Covenant: net debt / EBITDA must stay below")
    for c in C: put(m, f"{c}26", 5.5, BLUE, MULT)
    put(m, "A27", "  Covenant test")
    for c in C:
        put(m, f"{c}27", f'=IF({c}24<{c}26,"PASS","BREACH")', BLACK)
        m[f"{c}27"].fill = WARNFILL

    # fix row 21 reference used by row 13 of the following year
    for c in C[1:]:
        p_ = C[C.index(c) - 1]
        m[f"{c}13"].value = f"={p_}21"

    # ---- Returns ---------------------------------------------------------
    rt = wb.create_sheet("Returns")
    rt.column_dimensions["A"].width = 48
    for col in "BCDEFG": rt.column_dimensions[col].width = 13
    put(rt, "A1", "Sponsor returns", TITLE)
    put(rt, "A3", "EXIT", BOLD, fill=SECFILL)
    put(rt, "A4", "Exit year EBITDA");        put(rt, "B4", "=Model!F6", GREEN, CUR)
    put(rt, "A5", "Exit multiple");           put(rt, "B5", "=Assumptions!B13", GREEN, MULT)
    put(rt, "A6", "Exit enterprise value");   put(rt, "B6", "=B4*B5", BLACK, CUR)
    put(rt, "A7", "Less: net debt at exit");  put(rt, "B7", "=-Model!F23", BLACK, CUR)
    put(rt, "A8", "Equity value at exit", BOLD); put(rt, "B8", "=B6+B7", BOLD, CUR)

    put(rt, "A10", "CASH FLOWS TO THE SPONSOR", BOLD, fill=SECFILL)
    col_header(rt, 10, ["Close", "FY27", "FY28", "FY29", "FY30", "FY31"], start_col=2)
    put(rt, "A11", "Equity in / (out)")
    put(rt, "B11", "=-'Sources and Uses'!B12", BLACK, CUR)
    for col in "CDEF": put(rt, f"{col}11", 0, BLUE, CUR)
    put(rt, "G11", "=B8", BLACK, CUR)
    put(rt, "A13", "Money multiple (MoM)", BOLD)
    put(rt, "B13", "=B8/'Sources and Uses'!B12", BOLD, MULT); rt["B13"].fill = WARNFILL
    put(rt, "A14", "Internal rate of return (IRR)", BOLD)
    put(rt, "B14", "=IRR(B11:G11)", BOLD, PCT); rt["B14"].fill = WARNFILL
    put(rt, "A15", "  IRR cross-check from the multiple")
    put(rt, "B15", "=B13^(1/Assumptions!B14)-1", BLACK, PCT)
    put(rt, "C15", "should equal the IRR above when there are no interim cash flows", NOTE)

    put(rt, "A17", "WHERE THE MONEY CAME FROM", BOLD, fill=SECFILL)
    put(rt, "A18", "Value created for the sponsor")
    put(rt, "B18", "=B8-'Sources and Uses'!B12", BLACK, CUR)
    put(rt, "A19", "  EBITDA growth, held at the entry multiple")
    put(rt, "B19", "=(B4-Assumptions!B5)*Assumptions!B6", BLACK, CUR)
    put(rt, "C19", "=B19/$B$18", BLACK, PCT)
    put(rt, "D19", "this is operational work - it is the part you control", NOTE)
    put(rt, "A20", "  Multiple expansion, on exit EBITDA")
    put(rt, "B20", "=(B5-Assumptions!B6)*B4", BLACK, CUR)
    put(rt, "C20", "=B20/$B$18", BLACK, PCT)
    put(rt, "D20", "this is the market being kind to you - it is luck", NOTE)
    put(rt, "A21", "  Debt paydown and cash build")
    put(rt, "B21", "=Assumptions!B22-Model!F23", BLACK, CUR)
    put(rt, "C21", "=B21/$B$18", BLACK, PCT)
    put(rt, "D21", "the company's own cash flow repaying the borrowing", NOTE)
    put(rt, "A22", "  Fees and other")
    put(rt, "B22", "=B18-B19-B20-B21", BLACK, CUR)
    put(rt, "C22", "=B22/$B$18", BLACK, PCT)
    put(rt, "A23", "Total", BOLD)
    put(rt, "B23", "=SUM(B19:B22)", BOLD, CUR); put(rt, "C23", "=SUM(C19:C22)", BOLD, PCT)
    put(rt, "A24", "CHECK: total less value created", BOLD)
    put(rt, "B24", "=B23-B18", BOLD, CUR); rt["B24"].fill = WARNFILL

    put(rt, "A27", "THE PAPER LBO  - practise this until you can do it out loud in five minutes",
        BOLD, fill=SECFILL)
    paper = [
        ("Entry EBITDA", 100, CUR), ("Entry multiple", 10.0, MULT),
        ("Entry enterprise value", "=B28*B29", CUR), ("Debt (x EBITDA)", 6.0, MULT),
        ("Debt raised", "=B28*B31", CUR), ("Sponsor equity", "=B30-B32", CUR),
        ("EBITDA in year 5", 150, CUR), ("Debt repaid over the hold", 300, CUR),
        ("Debt at exit", "=B32-B35", CUR), ("Exit multiple", 10.0, MULT),
        ("Exit enterprise value", "=B34*B37", CUR), ("Equity at exit", "=B38-B36", CUR),
        ("Money multiple", "=B39/B33", MULT), ("IRR (5 years)", "=B40^(1/5)-1", PCT),
    ]
    r = 28
    for label, val, fmt in paper:
        put(rt, f"A{r}", label)
        put(rt, f"B{r}", val, BLUE if not isinstance(val, str) else BLACK, fmt)
        r += 1
    put(rt, "A42", "Memorise the mapping: over five years 2.0x is about 15% IRR, 2.5x about 20%,", NOTE)
    put(rt, "A43", "3.0x about 25%, 4.0x about 32%. Interviewers expect it instantly.", NOTE)

    # ---- Sensitivity -----------------------------------------------------
    s = wb.create_sheet("Sensitivity")
    s.column_dimensions["A"].width = 22
    for col in "BCDEFG": s.column_dimensions[col].width = 12
    put(s, "A1", "Sensitivity - sponsor IRR", TITLE)
    put(s, "A2", "Entry multiple down the side, exit multiple across the top.", NOTE)
    put(s, "A3", "Debt raised is a multiple of EBITDA, so it does not move with the entry multiple -", NOTE)
    put(s, "A4", "only the size of the equity cheque does. That is why this grid is exact and not an", NOTE)
    put(s, "A5", "approximation.", NOTE)
    entries = [9.0, 10.0, 11.0, 12.0, 13.0]
    exits = [9.0, 10.0, 11.0, 12.0, 13.0]
    put(s, "A7", "Entry \\ Exit", HEAD, fill=HEADFILL, align="center")
    for j, x in enumerate(exits):
        put(s, f"{get_column_letter(2+j)}7", x, BLUE, MULT, align="center")
    for i, e in enumerate(entries):
        row = 8 + i
        put(s, f"A{row}", e, BLUE, MULT, align="center")
        for j in range(len(exits)):
            col = get_column_letter(2 + j)
            # sponsor cheque = entry EV x (1 + fee%) - new debt
            cheque = f"($A{row}*Assumptions!$B$5*(1+Assumptions!$B$8)-Assumptions!$B$22)"
            exitq  = f"({col}$7*Model!$F$6-Model!$F$23)"
            put(s, f"{col}{row}", f"=({exitq}/{cheque})^(1/Assumptions!$B$14)-1", BLACK, PCT)
    put(s, "A15", "Money multiple", BOLD)
    put(s, "A16", "Entry \\ Exit", HEAD, fill=HEADFILL, align="center")
    for j, x in enumerate(exits):
        put(s, f"{get_column_letter(2+j)}16", x, BLUE, MULT, align="center")
    for i, e in enumerate(entries):
        row = 17 + i
        put(s, f"A{row}", e, BLUE, MULT, align="center")
        for j in range(len(exits)):
            col = get_column_letter(2 + j)
            cheque = f"($A{row}*Assumptions!$B$5*(1+Assumptions!$B$8)-Assumptions!$B$22)"
            exitq  = f"({col}$16*Model!$F$6-Model!$F$23)"
            put(s, f"{col}{row}", f"={exitq}/{cheque}", BLACK, MULT)

    readme_sheet(wb, f"{COMPANY['name']} - leveraged buyout model", LEGEND + [
        "# What this is",
        "A five-year LBO: sources and uses, an operating and debt model with a cash sweep, sponsor",
        "returns with attribution, a paper-LBO drill, and IRR sensitivity.",
        "",
        "# Sheets",
        "Assumptions        what you pay, how much you borrow, what you sell for",
        "Sources and Uses   where the money comes from and where it goes. Must balance.",
        "Model              operating forecast, interest, cash sweep, covenant test",
        "Returns            IRR, money multiple, and where the value actually came from",
        "Sensitivity        IRR and money multiple across entry and exit multiples",
        "",
        "# Things worth noticing",
        "1. The offer premium is an OUTPUT (Assumptions!B26), not an input. You decide what you can",
        "   pay from the returns you need; the premium is what falls out. Sponsors think this way",
        "   round and it is worth internalising early.",
        "2. Interest is charged on OPENING debt. Charging it on average debt makes the sweep circular:",
        "   interest depends on debt, debt depends on cash flow, cash flow depends on interest.",
        "   Banking models solve that with iterative calculation. This one avoids it. Both are valid;",
        "   know why you chose yours.",
        "3. This company holds more cash than debt, so the deal is funded partly with its own balance",
        "   sheet. That is normal, and it is why the sponsor cheque is smaller than the purchase price.",
        "4. Look at the attribution block on the Returns sheet before you look at the IRR. A deal that",
        "   makes its money from multiple expansion made it from luck. A deal that makes it from EBITDA",
        "   growth and debt paydown made it from work. Investment committees ask about this split.",
        "5. The covenant row is not decoration. Run a downside case - cut EBITDA growth to zero - and",
        "   find the year it breaches. That year is the deliverable of a credit review.",
        "",
        "# The default deal does not work, and that is the point",
        "At 11x entry the sponsor IRR is about 9%. No private equity fund accepts that. The model is",
        "not broken - it is telling you that a 14% premium for a cash-rich, low-margin retailer does",
        "not clear a 20% hurdle. Open the Sensitivity sheet: you would need to enter at about 9x, which",
        "means offering close to the undisturbed share price, which means the board says no. That whole",
        "chain of reasoning is a real deal conversation, and you just had it with a spreadsheet.",
        "",
        "# Try this",
        "Set the exit multiple equal to the entry multiple (it already is). Now the returns come only",
        "from growth and deleveraging - no help from the market. If the IRR still clears 20%, the deal",
        "stands on its own. If it only works at a higher exit multiple, you are betting on the market,",
        "and you should say so out loud rather than burying it in a cell.",
    ])
    wb.save(path)
    return path


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "..", "assets", "models")
    os.makedirs(out, exist_ok=True)
    for fn, name in ((build_three_statement, "01_three_statement_model.xlsx"),
                     (build_dcf,             "02_dcf_and_comps.xlsx"),
                     (build_lbo,             "03_lbo_model.xlsx")):
        print("built", fn(os.path.join(out, name)))
    print("\nNow recalculate each file (open in Excel, or use the xlsx skill's recalc.py)")
    print("and confirm every CHECK cell reads zero before trusting anything.")
